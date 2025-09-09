import pandas as pd
from openpyxl import load_workbook
from load_PER import get_financial_series, get_quarterly_financial_series, API_KEY
import warnings
warnings.filterwarnings('ignore')

def add_financial_columns(excel_file_path, new_columns, output_file_path=None, use_quarterly=False):
    """
    Excel 파일의 각 시트에 재무 데이터를 추가하는 함수
    
    Args:
        excel_file_path (str): 기존 Excel 파일 경로
        new_columns (list): 추가할 컬럼 이름들의 리스트 (예: ['EPS', 'ROE', 'ROA'])
        output_file_path (str): 출력 파일 경로 (None이면 기존 파일 덮어쓰기)
        use_quarterly (bool): True면 분기별 누적 데이터 사용, False면 연간 데이터 사용
    """
    
    wb = load_workbook(excel_file_path)
    
    if output_file_path is None:
        output_file_path = excel_file_path
    
    print(f"Excel 파일 로드 완료: {excel_file_path}")
    print(f"시트 개수: {len(wb.sheetnames)}")
    
    for sheet_name in wb.sheetnames:
        print(f"\n처리 중: {sheet_name}")
        
        try:
            ws = wb[sheet_name]
            
            # 헤더 행 읽기
            headers = [cell.value for cell in ws[1]]
            print(f"  기존 컬럼: {headers}")
            
            # 데이터 행 수 확인
            data_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:  # 날짜가 있는 행만
                    data_rows.append(row)
            
            if not data_rows:
                print(f"  {sheet_name}: 데이터가 없습니다.")
                continue
                
            print(f"  데이터 행 수: {len(data_rows)}")
            
            # 재무 데이터 가져오기
            print(f"  재무 데이터 수집 중...")
            if use_quarterly:
                financial_df = get_quarterly_financial_series(API_KEY, sheet_name, 2015, 2025)
            else:
                financial_df = get_financial_series(API_KEY, sheet_name, 2015, 2025)
            
            if financial_df.empty:
                print(f"  {sheet_name}: 재무 데이터를 가져올 수 없습니다.")
                continue
            
            print(f"  재무 데이터 수집 완료: {len(financial_df)}년")
            print(f"  재무 지표: {list(financial_df.columns)}")
            
            # 기존 헤더에 새로운 컬럼 추가
            for col in new_columns:
                if col not in headers:
                    headers.append(col)
            
            # 헤더 행 업데이트
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # 각 데이터 행에 대해 재무 데이터 매칭
            for row_idx, row_data in enumerate(data_rows, 2):
                # 날짜에서 연도 추출 (YYYY.MM.DD 형식)
                date_str = str(row_data[0])
                if '.' in date_str:
                    year = int(date_str.split('.')[0])
                else:
                    year = None
                
                if year and year in financial_df.index:
                    financial_data = financial_df.loc[year]
                    
                    # 새로운 컬럼에 재무 데이터 추가
                    for col_idx, col in enumerate(new_columns, len(row_data) + 1):
                        if col in financial_data:
                            value = financial_data[col]
                            ws.cell(row=row_idx, column=col_idx, value=value)
                        else:
                            ws.cell(row=row_idx, column=col_idx, value=None)
                else:
                    # 해당 연도 데이터가 없으면 None으로 채우기
                    for col_idx, col in enumerate(new_columns, len(row_data) + 1):
                        ws.cell(row=row_idx, column=col_idx, value=None)
            
            print(f"  {sheet_name}: 재무 데이터 추가 완료")
            
        except Exception as e:
            print(f"  {sheet_name}: 오류 발생 - {str(e)}")
            continue
    
    # 파일 저장
    wb.save(output_file_path)
    print(f"\n파일 저장 완료: {output_file_path}")

def add_financial_indicators(excel_file_path, output_file_path=None, use_quarterly=False):
    """
    기본 재무 지표들을 추가하는 편의 함수
    """
    financial_columns = ['EPS', 'ROE', 'ROA']
    add_financial_columns(excel_file_path, financial_columns, output_file_path, use_quarterly)
    
    print(f"\n추가된 컬럼들:")
    print(f"- EPS: 주당순이익")
    print(f"- ROE: 자기자본이익률")
    print(f"- ROA: 총자산이익률")

if __name__ == "__main__":
    # 설정
    input_file = "ship_stock_prices.xlsx"
    output_file = "ship_stock_prices_with_financials.xlsx"
    
    print("=== 주식 데이터와 재무 데이터 통합 시작 ===")
    
    # 기본 재무 지표 추가 (연간 데이터)
    add_financial_indicators(input_file, output_file, use_quarterly=False)
    
    # 분기별 데이터로 추가하려면:
    # add_financial_indicators(input_file, output_file, use_quarterly=True)
    
    print(f"\n=== 통합 완료 ===")
    print(f"결과 파일: {output_file}")