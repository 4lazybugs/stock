import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import time
from datetime import date, datetime
import os

# headers 설정
headers = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/124.0.0.0 Safari/537.36"),
    "Referer": "https://finance.naver.com/",
    "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
}

session = requests.Session()
session.headers.update(headers)

# 조선 관련 주식 code 
codes = [
    '042660',  # 한화오션 (구 대우조선해양)
    '009540',  # HD한국조선해양
    '010140',  # 삼성중공업
    '010620',  # 현대미포조선
    '329180',  # 현대중공업
    '097230',  # HJ중공업 (구 한진중공업)
    '238490',  # 현대힘스
    '077970',  # STX엔진
    '267250',  # HD현대마린엔진
]

# 주가 저장할 액셀 초기화
wb = Workbook()
wb.remove(wb.active)

# 주가 불러오기 + 액셀에 저장
for code in codes:
    ws = wb.create_sheet(title=code)
    ws.append(["date","Close","Open","High","Low","Volume"])
    
    print(f"크롤링 시작: {code}")
    START_PAGE = 2
    MAX_PAGES = 5
    
    for page in range(START_PAGE, MAX_PAGES + 1):
        # ✅ 일별시세 페이지 (table.type2가 여기에 있음)
        url = f"https://finance.naver.com/item/sise_day.naver?code={code}&page={page}"
        print(f"  페이지 {page} 크롤링 중...")
        
        r = session.get(url, timeout=10)
        r.encoding = "euc-kr"
        soup = BeautifulSoup(r.text, "html.parser")

        # 표에서 데이터 추출
        table = soup.select_one("table.type2")
        if not table:
            break

        rows_with_data = 0
        for row in table.select("tr"):
            tds = [td.get_text(strip=True) for td in row.find_all("td")]
            if len(tds) == 7 and tds[0]:   # 실제 데이터 행만
                trade_date, close, _, open_, high, low, volume = tds # 전일비는 무시

                date = datetime.strptime(trade_date, "%Y.%m.%d")
                ws.append([
                    date,
                    int(close.replace(",","")),
                    int(open_.replace(",","")),
                    int(high.replace(",","")),
                    int(low.replace(",","")),
                    int(volume.replace(",",""))
                ])

                ws.cell(ws.max_row, 1).number_format = "yyyy-mm-dd"
                rows_with_data += 1
        
        # 데이터가 없으면 마지막 페이지
        if rows_with_data == 0:
            break
            
        print(f"  페이지 {page}에서 {rows_with_data}개 데이터 수집")
        page += 1
        time.sleep(0.5)  # 서버 부하 방지
    
    print(f"크롤링 완료: {code} (총 {page-1}페이지)")

os.makedirs('./stk_data', exist_ok=True)
wb.save('./stk_data/ship_stock_prices.xlsx')